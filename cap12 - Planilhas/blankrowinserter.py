#!/home/pdajgs/python_labs/py3.7/bin/python3

import openpyxl, sys

try:
    lineReference = int(sys.argv[1])
    lineInsert = int(sys.argv[2])
    fileExcel = sys.argv[3]
except ValueError as e:
    print('Os primeiros valores devem ser inteiros, conforme exemplo... 3 2 file.xlsx')

wb = openpyxl.load_workbook(fileExcel)
sheet = wb.active

wbWrite = openpyxl.Workbook()
sheetWrite = wbWrite.active

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        if row < lineReference:
            sheetWrite.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value
        else:
            sheetWrite.cell(row=(row + lineInsert), column=col).value = sheet.cell(row=row, column=col).value

wbWrite.save('xInsert.xlsx')
print('Done')