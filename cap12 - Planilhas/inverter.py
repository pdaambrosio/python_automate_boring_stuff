#!/home/pdajgs/python_labs/py3.7/bin/python3

import openpyxl, sys

fileExcel = sys.argv[1]

readWb = openpyxl.load_workbook(fileExcel)
readSheet = readWb.active

writeWb = openpyxl.Workbook()
writeSheet = writeWb.active

for readRow in range(1, readSheet.max_row + 1):
    for readCol in range(1, readSheet.max_column + 1):
        writeSheet.cell(row=readCol, column=readRow).value = readSheet.cell(row=readRow, column=readCol).value

writeWb.save('xInverter.xlsx')

print('Done!')