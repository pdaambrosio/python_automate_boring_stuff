#!/home/pdajgs/python_labs/py3.7/bin/python3

import openpyxl, sys
from openpyxl.styles import Font, NamedStyle

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'MultiplicationTable'

font1 = Font(bold=True)

num = sys.argv[1]

for i in range(1, int(num)+1):
    sheet['A' + str(i+1)] = i
    sheet.cell(row=1, column=i+1).value = i

    sheet['A' + str(i+1)].font = font1
    sheet.cell(row=1, column=i+1).font =font1

for j in range(1, int(num)+1):
    for k in range(1, int(num)+1):
        sheet.cell(row= j+1, column= k+1).value = j*k

wb.save('multiplicationTable.xlsx')
print('Done')