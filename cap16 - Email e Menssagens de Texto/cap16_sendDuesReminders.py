#!/home/pdajgs/python_labs/py3.7/bin/python3

import openpyxl, smtplib, sys

# abrir a planilha e obter o status do ultimo pagamento
planilha = openpyxl.load_workbook('./duesRecords.xlsx')
pastaTr = planilha.get_sheet_by_name('Sheet1')

colunaFinal = pastaTr.max_column
ultimoMes = pastaTr.cell(row=1, column=colunaFinal).value

devedores = {}

for i in range(2, pastaTr.max_row+1):
    pagamento = pastaTr.cell(row=i,column=colunaFinal).value
    if pagamento != 'paid':
        nome = pastaTr.cell(row=i, column=1).value
        email = pastaTr.cell(row=i, column=2).value
        devedores[nome] = email

# fazendo login no e-mail
smtpLogin = smtplib.SMTP('smtp.gmail.com', '587')
smtpLogin.ehlo()
smtpLogin.starttls()

smtpLogin.login('pda.ambrosio@gmail.com', sys.argv[1])
# smtpLogin.login('pda.ambrosio@gmail.com', input())

for nome, email in devedores.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make this as soon as possible. Thank you!" %(ultimoMes,nome,email)
    print('Sending email to %s' %email)
    sendmailStatus = smtpLogin.sendmail('pda.ambrosio@gmail.com',email,body)
    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' %(email,sendmailStatus))
smtpLogin.quit()